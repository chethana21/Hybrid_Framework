# getting num of row
# getting num of col
# read cell
# write
# color
import openpyxl
from openpyxl.styles import PatternFill

def GetRow(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_row

def GetCol(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_column

def ReadCell(xl_path, sheet, r,c):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.cell(r,c).value

def WriteCell(xl_path, sheet, r,c, data):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    sh.cell(r,c).value = data
    wb.save(xl_path)

def FillGreen(xl_path, sheet, r,c):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    green = PatternFill(start_color="21ad17", end_color="21ad17", fill_type="solid")
    sh.cell(r, c).fill = green
    wb.save(xl_path)

def FillRed(xl_path, sheet, r,c):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    red = PatternFill(start_color="fc1008Z", end_color="fc1008", fill_type="solid")
    sh.cell(r, c).fill = red
    wb.save(xl_path)
